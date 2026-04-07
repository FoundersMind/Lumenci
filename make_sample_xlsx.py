from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Claim Chart"

    headers = [
        "Patent Claim Element",
        "Accused Product Feature / Evidence",
        "AI Reasoning",
    ]
    ws.append(headers)

    ws.append(
        [
            "A temperature control device with a wireless communication module",
            'Acme Thermostat product page states: "WiFi-enabled smart thermostat connects to your home network"',
            "The Acme device has WiFi capability which satisfies the wireless communication module requirement",
        ]
    )
    ws.append(
        [
            "A motion sensor for detecting occupancy",
            'Acme technical specifications document shows: "Built-in motion sensor detects when people are home"',
            "Motion sensor explicitly mentioned in specs directly maps to the claim element for occupancy detection",
        ]
    )
    ws.append(
        [
            "Machine learning algorithm that learns user temperature preferences over time",
            'Acme marketing materials claim: "Auto-Schedule learns your preferred temperatures"',
            "The learning behavior described suggests ML algorithm, though technical implementation details are not disclosed. May need stronger evidence.",
        ]
    )

    fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 42

    ws.row_dimensions[1].height = 22
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 60

    wb.save("sample_claim_chart.xlsx")
    print("WROTE sample_claim_chart.xlsx")


if __name__ == "__main__":
    main()

